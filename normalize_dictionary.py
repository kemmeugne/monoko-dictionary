"""
=============================================================
  Dialect Dictionary Normalizer
=============================================================
  Transforms raw Excel files (collected from field workers)
  into a normalized database-ready format with 3 tables:
  Words, Senses, and Examples.

  Usage:
    python normalize_dictionary.py

  Place your Excel files in an "input" folder next to this
  script. Normalized output will be saved to an "output" folder.

  Expected Excel format (per row):
    Col B: French word
    Col C: Dialect word (sense 1)
    Col D: Example sentence in dialect
    Col E: French translation of example
    Col F: Dialect word (sense 2)
    Col G: Example sentence (sense 2)
    Col H: French translation (sense 2)
    ... and so on for senses 3, 4, etc.
=============================================================
"""

import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── CONFIGURATION ──────────────────────────────────────────
# Change this to the name of the dialect you're processing.
# It will appear in the "language" column of the Words table.
LANGUAGE_NAME = "Lingala"

# How many senses (meanings) does your Excel support at most?
# Your current format has 4 (columns C-E, F-H, I-K, L-N).
# If some files have more, increase this number.
MAX_SENSES = 4

# First row of actual data (0-indexed). Your files have
# 2 header rows (row 0 = "Sens n°1/2/3...", row 1 = column names),
# so data starts at row index 2.
DATA_START_ROW = 2

# Column index (0-based) where the French word is.
FRENCH_WORD_COL = 1

# Column index (0-based) where sense 1 starts.
# Sense 1 = cols 2,3,4 | Sense 2 = cols 5,6,7 | etc.
FIRST_SENSE_COL = 2
COLS_PER_SENSE = 3  # (dialect word, sentence, translation)
# ───────────────────────────────────────────────────────────


def parse_excel(filepath):
    """Parse one Excel file and return lists of words, senses, examples."""
    df = pd.read_excel(filepath, header=None)
    words = []
    senses = []
    examples = []

    for i in range(DATA_START_ROW, len(df)):
        row = df.iloc[i]
        french_word = row[FRENCH_WORD_COL]

        if pd.isna(french_word) or str(french_word).strip() == "":
            continue

        french_word = str(french_word).strip()
        letter = french_word[0].upper()

        word_entry = {
            "french_word": french_word,
            "letter": letter,
            "language": LANGUAGE_NAME,
        }
        word_senses = []
        word_examples = []

        sense_num = 0
        for s in range(MAX_SENSES):
            col_start = FIRST_SENSE_COL + (s * COLS_PER_SENSE)

            dialect_word = row[col_start] if col_start < len(row) else None
            sentence_dialect = row[col_start + 1] if col_start + 1 < len(row) else None
            sentence_french = row[col_start + 2] if col_start + 2 < len(row) else None

            if pd.notna(dialect_word) and str(dialect_word).strip() != "":
                sense_num += 1
                sense_entry = {
                    "sense_number": sense_num,
                    "dialect_word": str(dialect_word).strip(),
                }
                word_senses.append(sense_entry)

                if pd.notna(sentence_dialect) and pd.notna(sentence_french):
                    example_entry = {
                        "sense_number": sense_num,
                        "sentence_dialect": str(sentence_dialect).strip(),
                        "sentence_french": str(sentence_french).strip(),
                    }
                    word_examples.append(example_entry)

        words.append(word_entry)
        senses.append(word_senses)
        examples.append(word_examples)

    return words, senses, examples


def build_normalized_excel(all_words, all_senses, all_examples, output_path):
    """Build a single normalized Excel workbook from parsed data."""

    # Flatten and assign IDs
    words_flat = []
    senses_flat = []
    examples_flat = []

    word_id = 0
    sense_id = 0
    example_id = 0

    for w, w_senses, w_examples in zip(all_words, all_senses, all_examples):
        word_id += 1
        words_flat.append({
            "word_id": word_id,
            "french_word": w["french_word"],
            "letter": w["letter"],
            "language": w["language"],
        })

        sense_id_map = {}  # sense_number -> sense_id for this word
        for s in w_senses:
            sense_id += 1
            sense_id_map[s["sense_number"]] = sense_id
            senses_flat.append({
                "sense_id": sense_id,
                "word_id": word_id,
                "sense_number": s["sense_number"],
                "dialect_word": s["dialect_word"],
            })

        for ex in w_examples:
            example_id += 1
            examples_flat.append({
                "example_id": example_id,
                "sense_id": sense_id_map[ex["sense_number"]],
                "word_id": word_id,
                "sentence_dialect": ex["sentence_dialect"],
                "sentence_french": ex["sentence_french"],
            })

    # ── Create Excel workbook ──
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="2B579A")
    data_font = Font(size=11, name="Arial")
    alt_fill = PatternFill("solid", fgColor="F2F6FC")
    border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    def make_sheet(ws, title, headers, widths, data, keys):
        ws.title = title
        ws.freeze_panes = "A2"
        for col_idx, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.row_dimensions[1].height = 25

        for row_idx, item in enumerate(data, 2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, key in enumerate(keys, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=item[key])
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if fill:
                    cell.fill = fill

        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(data) + 1}"

    # Words sheet
    make_sheet(
        wb.active, "Words",
        ["word_id", "french_word", "letter", "language"],
        [12, 25, 10, 15],
        words_flat,
        ["word_id", "french_word", "letter", "language"],
    )

    # Senses sheet
    make_sheet(
        wb.create_sheet(), "Senses",
        ["sense_id", "word_id", "sense_number", "dialect_word"],
        [12, 12, 15, 35],
        senses_flat,
        ["sense_id", "word_id", "sense_number", "dialect_word"],
    )

    # Examples sheet
    make_sheet(
        wb.create_sheet(), "Examples",
        ["example_id", "sense_id", "word_id", "sentence_dialect", "sentence_french"],
        [12, 12, 12, 50, 50],
        examples_flat,
        ["example_id", "sense_id", "word_id", "sentence_dialect", "sentence_french"],
    )

    wb.save(output_path)
    return len(words_flat), len(senses_flat), len(examples_flat)


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_dir = os.path.join(script_dir, "input")
    output_dir = os.path.join(script_dir, "output")

    # Create folders if they don't exist
    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    # Find all Excel files in input folder
    excel_files = [
        f for f in os.listdir(input_dir)
        if f.endswith((".xlsx", ".xls")) and not f.startswith("~")
    ]

    if not excel_files:
        print("=" * 50)
        print("No Excel files found!")
        print(f"Place your .xlsx files in: {input_dir}")
        print("=" * 50)
        sys.exit(1)

    excel_files.sort()
    print(f"\nFound {len(excel_files)} file(s) to process:")
    for f in excel_files:
        print(f"  • {f}")
    print()

    # Parse all files
    all_words = []
    all_senses = []
    all_examples = []

    for filename in excel_files:
        filepath = os.path.join(input_dir, filename)
        print(f"Processing {filename}...", end=" ")
        try:
            words, senses, examples = parse_excel(filepath)
            all_words.extend(words)
            all_senses.extend(senses)
            all_examples.extend(examples)
            print(f"OK ({len(words)} words)")
        except Exception as e:
            print(f"ERROR: {e}")

    # Build output
    output_path = os.path.join(output_dir, f"dictionary_{LANGUAGE_NAME.lower()}_normalized.xlsx")
    n_words, n_senses, n_examples = build_normalized_excel(
        all_words, all_senses, all_examples, output_path
    )

    print()
    print("=" * 50)
    print("  DONE!")
    print("=" * 50)
    print(f"  Words:    {n_words}")
    print(f"  Senses:   {n_senses}")
    print(f"  Examples:  {n_examples}")
    print(f"  Output:    {output_path}")
    print("=" * 50)


if __name__ == "__main__":
    main()
