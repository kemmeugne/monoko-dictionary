#!/usr/bin/env python3
"""
Build and optionally validate a manifest for Lingala dictionary audio files.

The audio filenames map back to the original workbook cells, e.g.:
  P.C39.mp3  -> workbook P.xlsx, cell C39
  P.D63.mp3  -> workbook P.xlsx, cell D63

Supported audio-bearing columns:
  C/F/I/L -> dialect word for sense 1/2/3/4
  D/G/J/M -> dialect example sentence for sense 1/2/3/4

Usage examples:
  python3 lingala_audio_manifest.py
  python3 lingala_audio_manifest.py --validate-supabase

Required env vars for --validate-supabase:
  SUPABASE_URL
  SUPABASE_SERVICE_KEY
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import load_workbook

try:
    from supabase import create_client
except ImportError:
    create_client = None


DEFAULT_AUDIO_ROOT = Path(
    "/Users/anthonykemmeugne/Documents/App dialectes/Lingala/Audios Lingala Finaux (01.02.2026)"
)
DEFAULT_WORKBOOK_ROOT = Path(
    "/Users/anthonykemmeugne/Documents/App dialectes/Lingala/Tableau Lingala FINAL (01.02.2026)/Tableau Dictionnaire Lingala"
)
DEFAULT_OUTPUT_DIR = Path("artifacts/lingala_audio")

FRENCH_COL = "B"
SENSE_COLS = {
    "C": {"sense_number": 1, "kind": "sense_word", "dialect_col": "C", "example_fr_col": "E"},
    "D": {"sense_number": 1, "kind": "example_sentence", "dialect_col": "C", "example_fr_col": "E"},
    "F": {"sense_number": 2, "kind": "sense_word", "dialect_col": "F", "example_fr_col": "H"},
    "G": {"sense_number": 2, "kind": "example_sentence", "dialect_col": "F", "example_fr_col": "H"},
    "I": {"sense_number": 3, "kind": "sense_word", "dialect_col": "I", "example_fr_col": "K"},
    "J": {"sense_number": 3, "kind": "example_sentence", "dialect_col": "I", "example_fr_col": "K"},
    "L": {"sense_number": 4, "kind": "sense_word", "dialect_col": "L", "example_fr_col": "N"},
    "M": {"sense_number": 4, "kind": "example_sentence", "dialect_col": "L", "example_fr_col": "N"},
}
ROW_RE = re.compile(r"^(?P<letter>[A-Z])\.(?P<column>[A-Z]+)(?P<row>\d+)\.(?P<ext>mp3|wav|m4a|ogg)$", re.IGNORECASE)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")
    return " ".join(text.split()).strip()


@dataclass
class ManifestRow:
    audio_path: str
    audio_file: str
    workbook_letter: str
    workbook_path: str
    source_cell: str
    row_number: int
    source_column: str
    target_type: str
    target_table: str
    sense_number: int
    french_word: str
    dialect_word: str
    sentence_dialect: str
    sentence_french: str
    object_key: str
    public_url: str
    db_match_status: str = "unvalidated"
    db_word_id: str = ""
    db_sense_id: str = ""
    db_example_id: str = ""
    db_match_note: str = ""


def iter_audio_files(audio_root: Path) -> Iterable[Path]:
    for path in sorted(audio_root.rglob("*")):
        if path.is_file() and ROW_RE.match(path.name):
            yield path


def get_sheet(workbook_root: Path, letter: str, cache: Dict[str, tuple]):
    if letter in cache:
        return cache[letter]
    workbook_path = workbook_root / f"{letter}.xlsx"
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found for letter {letter}: {workbook_path}")
    ws = load_workbook(workbook_path, data_only=True, read_only=True).active
    cache[letter] = (workbook_path, ws)
    return cache[letter]


def build_manifest_row(
    audio_path: Path,
    workbook_root: Path,
    r2_prefix: str,
    public_base_url: str,
    workbook_cache: Dict[str, tuple],
) -> ManifestRow:
    match = ROW_RE.match(audio_path.name)
    if not match:
        raise ValueError(f"Unsupported audio filename: {audio_path.name}")

    letter = match.group("letter").upper()
    column = match.group("column").upper()
    row_number = int(match.group("row"))
    if column not in SENSE_COLS:
        raise ValueError(f"Column {column} is not a Lingala audio target: {audio_path.name}")

    workbook_path, ws = get_sheet(workbook_root, letter, workbook_cache)
    french_word = clean_text(ws[f"{FRENCH_COL}{row_number}"].value)
    mapping = SENSE_COLS[column]
    dialect_word = clean_text(ws[f"{mapping['dialect_col']}{row_number}"].value)
    sentence_dialect = clean_text(ws[f"{column}{row_number}"].value) if mapping["kind"] == "example_sentence" else ""
    sentence_french = clean_text(ws[f"{mapping['example_fr_col']}{row_number}"].value)

    if not french_word:
        raise ValueError(f"French source cell is empty for {audio_path.name} -> {FRENCH_COL}{row_number}")
    if not dialect_word:
        raise ValueError(f"Dialect word cell is empty for {audio_path.name}")
    if mapping["kind"] == "example_sentence" and not sentence_dialect:
        raise ValueError(f"Dialect example cell is empty for {audio_path.name}")

    target_table = "senses" if mapping["kind"] == "sense_word" else "examples"
    object_key = f"{r2_prefix}/{target_table}/{letter}/{audio_path.name}"
    public_url = f"{public_base_url.rstrip('/')}/{object_key}"

    return ManifestRow(
        audio_path=str(audio_path),
        audio_file=audio_path.name,
        workbook_letter=letter,
        workbook_path=str(workbook_path),
        source_cell=f"{column}{row_number}",
        row_number=row_number,
        source_column=column,
        target_type=mapping["kind"],
        target_table=target_table,
        sense_number=mapping["sense_number"],
        french_word=french_word,
        dialect_word=dialect_word,
        sentence_dialect=sentence_dialect,
        sentence_french=sentence_french,
        object_key=object_key,
        public_url=public_url,
    )


def fetch_lingala_records(supabase_url: str, service_key: str, language_id: int) -> Tuple[Dict[Tuple[str, int, str], dict], Dict[Tuple[str, int, str, str], dict]]:
    if create_client is None:
        raise RuntimeError("supabase package is not installed")

    client = create_client(supabase_url, service_key)
    sense_lookup: Dict[Tuple[str, int, str], dict] = {}
    example_lookup: Dict[Tuple[str, int, str, str], dict] = {}

    page_size = 1000
    offset = 0
    while True:
        response = (
            client.table("words")
            .select("id,french_word,letter,senses(id,sense_number,dialect_word,examples(id,sentence_dialect,sentence_french))")
            .eq("language_id", language_id)
            .range(offset, offset + page_size - 1)
            .execute()
        )
        rows = response.data or []
        if not rows:
            break

        for word in rows:
            french_word = clean_text(word.get("french_word"))
            for sense in word.get("senses") or []:
                sense_key = (
                    french_word.casefold(),
                    int(sense["sense_number"]),
                    clean_text(sense.get("dialect_word")).casefold(),
                )
                sense_lookup[sense_key] = {
                    "word_id": word["id"],
                    "sense_id": sense["id"],
                }
                for example in sense.get("examples") or []:
                    example_key = (
                        french_word.casefold(),
                        int(sense["sense_number"]),
                        clean_text(example.get("sentence_dialect")).casefold(),
                        clean_text(example.get("sentence_french")).casefold(),
                    )
                    example_lookup[example_key] = {
                        "word_id": word["id"],
                        "sense_id": sense["id"],
                        "example_id": example["id"],
                    }

        if len(rows) < page_size:
            break
        offset += page_size

    return sense_lookup, example_lookup


def validate_rows(rows: List[ManifestRow], supabase_url: str, service_key: str, language_id: int) -> None:
    sense_lookup, example_lookup = fetch_lingala_records(supabase_url, service_key, language_id)

    for row in rows:
        french_key = row.french_word.casefold()
        if row.target_table == "senses":
            lookup_key = (french_key, row.sense_number, row.dialect_word.casefold())
            match = sense_lookup.get(lookup_key)
            if not match:
                row.db_match_status = "missing"
                row.db_match_note = "No exact sense match in Supabase"
                continue
            row.db_match_status = "matched"
            row.db_word_id = str(match["word_id"])
            row.db_sense_id = str(match["sense_id"])
            row.db_match_note = "Exact sense match"
            continue

        lookup_key = (
            french_key,
            row.sense_number,
            row.sentence_dialect.casefold(),
            row.sentence_french.casefold(),
        )
        match = example_lookup.get(lookup_key)
        if not match:
            row.db_match_status = "missing"
            row.db_match_note = "No exact example match in Supabase"
            continue
        row.db_match_status = "matched"
        row.db_word_id = str(match["word_id"])
        row.db_sense_id = str(match["sense_id"])
        row.db_example_id = str(match["example_id"])
        row.db_match_note = "Exact example match"


def write_outputs(rows: List[ManifestRow], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "lingala_audio_manifest.json"
    csv_path = output_dir / "lingala_audio_manifest.csv"
    summary_path = output_dir / "lingala_audio_summary.json"

    dict_rows = [asdict(row) for row in rows]
    with json_path.open("w", encoding="utf-8") as handle:
        json.dump(dict_rows, handle, ensure_ascii=False, indent=2)

    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(dict_rows[0].keys()) if dict_rows else [])
        if dict_rows:
            writer.writeheader()
            writer.writerows(dict_rows)

    summary = {
        "total_audio_files": len(rows),
        "by_target_table": {
            "senses": sum(1 for row in rows if row.target_table == "senses"),
            "examples": sum(1 for row in rows if row.target_table == "examples"),
        },
        "by_match_status": {},
    }
    for row in rows:
        summary["by_match_status"][row.db_match_status] = summary["by_match_status"].get(row.db_match_status, 0) + 1

    with summary_path.open("w", encoding="utf-8") as handle:
        json.dump(summary, handle, ensure_ascii=False, indent=2)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the Lingala audio manifest")
    parser.add_argument("--audio-root", type=Path, default=DEFAULT_AUDIO_ROOT)
    parser.add_argument("--workbook-root", type=Path, default=DEFAULT_WORKBOOK_ROOT)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--r2-prefix", default="Lingala")
    parser.add_argument("--public-base-url", default="https://pub-78d23bf07fce46b3adc19df91148ffb8.r2.dev")
    parser.add_argument("--validate-supabase", action="store_true")
    parser.add_argument("--language-id", type=int, default=1)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows: List[ManifestRow] = []
    errors: List[str] = []
    workbook_cache: Dict[str, tuple] = {}

    for audio_path in iter_audio_files(args.audio_root):
        try:
            rows.append(
                build_manifest_row(
                    audio_path,
                    args.workbook_root,
                    args.r2_prefix,
                    args.public_base_url,
                    workbook_cache,
                )
            )
        except Exception as exc:
            errors.append(f"{audio_path}: {exc}")

    if args.validate_supabase:
        supabase_url = os.environ.get("SUPABASE_URL")
        service_key = os.environ.get("SUPABASE_SERVICE_KEY")
        if not supabase_url or not service_key:
            raise RuntimeError("SUPABASE_URL and SUPABASE_SERVICE_KEY are required for --validate-supabase")
        validate_rows(rows, supabase_url, service_key, args.language_id)

    write_outputs(rows, args.output_dir)

    print(f"Manifest rows: {len(rows)}")
    if args.validate_supabase:
        matched = sum(1 for row in rows if row.db_match_status == "matched")
        print(f"Matched rows:  {matched}")
        print(f"Missing rows:  {sum(1 for row in rows if row.db_match_status == 'missing')}")
    if errors:
        error_path = args.output_dir / "lingala_audio_errors.txt"
        error_path.parent.mkdir(parents=True, exist_ok=True)
        error_path.write_text("\n".join(errors), encoding="utf-8")
        print(f"Errors:        {len(errors)} (see {error_path})")


if __name__ == "__main__":
    main()
